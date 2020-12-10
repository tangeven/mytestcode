# -*- coding:UTF-8 -*-

import requests
import json
import pandas as pd
from datetime import datetime
import os
from openpyxl.styles import *
from pprint import pprint
import openpyxl
from datetime import datetime, timedelta


class Wechat():
    def __init__(self):
        self.appid = 'wx254a4d6e0a1c0575'
        self.secureid = '4ab381f063eb9758c2ad05217177105c'
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.75 Safari/537.36"}
        self.token_url = 'https://api.weixin.qq.com/cgi-bin/token?grant_type=client_credential&appid={}&secret={}'
        self.begin_date = '2020-10-28'
        self.end_date = '2020-12-02'


    def get_autotoken(self):
        url = self.token_url.format(self.appid, self.secureid)
        response_ = requests.get(url, self.headers)
        json_data = response_.content.decode('utf-8')
        dict_data = json.loads(json_data)
        print(dict_data)
        if 'errcode' not in dict_data:
            access_token = dict_data['access_token']
            expire_time = dict_data['expires_in']
            print(access_token)
            return access_token


    def get_targetData(self, access_token):
        # print(access_token)

        begin_date = self.begin_date
        tmp_date = (datetime.strptime(begin_date, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d")
        end_date = self.end_date
        res_new = {}
        res_cancel = {}

        des_url_increase = 'https://api.weixin.qq.com/datacube/getusersummary?access_token={}'.format(access_token)
        des_url_sum = 'https://api.weixin.qq.com/datacube/getusercumulate?access_token={}'.format(access_token)
        while end_date > tmp_date:
            print(tmp_date)
            post_data = {"begin_date":self.begin_date, "end_date":tmp_date}
            post_data = json.dumps(post_data)
            # 总共的人数
            resp_sum = requests.post(des_url_sum, data=post_data)
            json_data_des_sum = resp_sum.content.decode('utf-8')
            des_dict_sum = json.loads(json_data_des_sum)
            # print(json_data_des_sum)

            # 新增的人数
            resp_increase = requests.post(des_url_increase, data=post_data)
            json_data_des_increase = resp_increase.content.decode('utf-8')
            des_dict_increase = json.loads(json_data_des_increase)

            res_increase = des_dict_increase['list']

            for i in res_increase:
                date = i['ref_date']

                if date not in res_new.keys():
                    res_new[date] = i['new_user']
                else:
                    res_new[date] += i['new_user']

            for j in res_increase:
                date = j['ref_date']
                if date not in res_cancel.keys():
                    res_cancel[date] = j['cancel_user']
                else:
                    res_cancel[date] += j['cancel_user']

            self.begin_date = tmp_date
            tmp_date = (datetime.strptime(tmp_date, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d")



        return res_new,res_cancel


    def write_excel(self,new,cancel):
        out_file_name = 'E:/留存.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('data',0)

        new_dict = new
        cancel_dict = cancel
        ws.cell(row=1, column=1, value='时间')
        ws.cell(row=1, column=2, value='新增')
        ws.cell(row=1, column=3, value='取消')
        ws.cell(row=1, column=4, value='净增')
        # ws.cell(row=1, column=5, value='')
        row_num_01 = 1
        row_num_02 = 1

        for item in new_dict.items():
            time = item[0]
            new_num = item[1]
            ws.cell(row_num_01+1,1).value = time
            ws.cell(row_num_01+1,2).value = new_num
            row_num_01 += 1

        for item in cancel_dict.items():
            cancel_num = item[1]
            ws.cell(row_num_02+1,3).value = cancel_num
            row_num_02 += 1

        wb.save(out_file_name)










    def run(self):
        access_token = self.get_autotoken()
        new,cancel = self.get_targetData(access_token)
        self.write_excel(new,cancel)

if __name__ == "__main__":
    wechat = Wechat()
    wechat.run()